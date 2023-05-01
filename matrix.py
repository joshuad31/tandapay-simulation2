import itertools
import os
import time
from datetime import datetime
from openpyxl import load_workbook

from utils.logger import logger
from tandapay import TandaPaySimulatorV2

# =============================  Test Data  =========================================
EV_LIST = [
    [60, 68, 70, 85],               # EV1
    [1000],                         # EV2
    [.40, .55, .70],                # EV3
    [.24, .26, .27, .28, .30],      # EV4
    [.10, .20],                     # EV5
    [.70, 1.00],                    # EV6
    [2, 3],                         # EV7
    [1],                            # EV8
    [.3333]                         # EV9
]
PV_LIST = [
    [.10],                          # PV1
    [.02],                          # PV2
    [.7],                           # PV3
    [.08, .12, .16, .20],           # PV4
    [.77, .83, .89, .95],           # PV5
    [.03]                           # PV6
]

# ===================================================================================

_cur_dir = os.path.dirname(os.path.realpath(__file__))
os.makedirs(os.path.join(_cur_dir, 'result'), exist_ok=True)


def create_matrix_result(ev_list, pv_list, target_dir=None, status_signal=None, finished_signal=None):

    dataset = [{'ev': list(d[:9]), 'pv': list(d[9:])} for d in list(itertools.product(*(ev_list + pv_list)))]

    s_time = time.time()
    results = []
    workbook = load_workbook(os.path.join(_cur_dir, 'databases', '3 Matrix Database.xlsx'))
    sh_map = workbook[workbook.sheetnames[0]]
    sh_log = workbook[workbook.sheetnames[1]]

    for i, d in enumerate(dataset):
        logger.debug(f"========== Processing {i}")
        if status_signal is not None and i % 10 == 0:
            status_signal.emit(i / len(dataset) * 100)
        ev = d['ev']
        pv = d['pv']
        sim = TandaPaySimulatorV2(ev=ev, pv=pv, matrix=True)
        r = sim.start_simulation()
        collapsed = 1 if r[1] / r[0] < .5 else 0
        results.append({'ev': ev, 'pv': pv, 'c': collapsed, 'result': r})
        sh_map.cell(i + 5, 1).value = i + 1

        offset = 2
        for k, evs in enumerate(ev_list):
            if len(evs) == 1:
                continue
            sh_map.cell(i + 5, offset + evs.index(ev[k])).value = collapsed
            offset += len(evs)
        for k, pvs in enumerate(pv_list):
            if len(pvs) == 1:
                continue
            sh_map.cell(i + 5, offset + pvs.index(pv[k])).value = collapsed
            offset += len(pvs)

        sh_log.cell(i + 2, 1).value = i + 1
        for j in range(9):
            ratio = 100 if (1 < j < 6 or j == 8) else 1     # EV3 ~ EV6, EV9 are percentage values
            sh_log.cell(i + 2, 2 + j).value = ev[j] * ratio
        for j in range(6):          # All PV values are percentages
            sh_log.cell(i + 2, 11 + j).value = pv[j] * 100
        for j, v in enumerate(r):
            sh_log.cell(i + 2, 17 + j).value = v

    offset = 2
    for i, evs in enumerate(ev_list):
        if len(evs) == 1:
            continue
        for j, ev in enumerate(evs):
            sh_map.cell(1, offset + j).value = f"EV{i + 1} = {ev * (100 if (1 < i < 6 or i == 8) else 1)}"
            sh_map.cell(2, offset + j).value = len([r for r in results if r['ev'][i] == ev])
            sh_map.cell(3, offset + j).value = len([r for r in results if r['ev'][i] == ev and r['c']])
            sh_map.cell(4, offset + j).value = \
                round(sh_map.cell(3, offset + j).value / sh_map.cell(2, offset + j).value * 100, 2)
        offset += len(evs)
    for i, pvs in enumerate(pv_list):
        if len(pvs) == 1:
            continue
        for j, pv in enumerate(pvs):
            sh_map.cell(1, offset + j).value = f"PV{i + 1} = {pv * 100}"
            sh_map.cell(2, offset + j).value = len([r for r in results if r['pv'][i] == pv])
            sh_map.cell(3, offset + j).value = len([r for r in results if r['pv'][i] == pv and r['c']])
            sh_map.cell(4, offset + j).value = \
                round(sh_map.cell(3, offset + j).value / sh_map.cell(2, offset + j).value * 100, 2)
        offset += len(pvs)

    if target_dir is None:
        target_dir = os.path.join(_cur_dir, 'result')
    result_file = os.path.join(target_dir, f"Matrix_{datetime.now().strftime('%m_%d_%Y__%H_%M_%S')}.xlsx")
    workbook.save(result_file)
    workbook.close()

    logger.debug(f"Saved to {result_file}, elapsed: {time.time() - s_time}")
    if finished_signal is not None:
        finished_signal.emit()
    return result_file


if __name__ == '__main__':

    create_matrix_result(ev_list=EV_LIST, pv_list=PV_LIST)
