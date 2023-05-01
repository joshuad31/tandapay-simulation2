import copy
import os
import shutil
import time
from datetime import datetime
import random
from openpyxl import load_workbook
from settings import RESULT_DIR, DB_DIR, FILE_SYSTEM, FILE_USER
from utils.logger import logger


class TandaPaySimulatorV2(object):

    def __init__(self, ev=None, pv=None, matrix=False):
        if len(ev) != 10:
            ev.append(ev[1] * 0.025 * ev[0])
        self.ev = copy.deepcopy(ev)
        self.pv = pv
        self.wb = {
            'user': load_workbook(os.path.join(DB_DIR, FILE_USER)),
            'system': load_workbook(os.path.join(DB_DIR, FILE_SYSTEM)),
        }
        self.sh_user = self.wb['user'].active
        self.sh_system = self.wb['system'].active

        self.excel_files = {}
        self.counter = 0
        self.matrix = matrix

        self.sy_rec_p = [self.sh_system.cell(2, i + 2) for i in range(22)]
        self.sy_rec_f = [self.sh_system.cell(3, i + 2) for i in range(22)]
        self.sy_rec_r = [self.sh_system.cell(4, i + 2) for i in range(22)]

    def _init_system_sheet(self, target_dir):
        for i in range(2):
            self.sh_system.cell(i + 2, 3).value = self.ev[0]
            self.sh_system.cell(i + 2, 4).value = self.ev[9] / self.ev[0]
            for k in range(5, 23):
                self.sh_system.cell(i + 2, k).value = 0 if k != 18 else 'no'
            self.sh_system.cell(i + 2, 21).value = self.ev[9] / self.ev[0]
        for i in range(3, 30):
            for k in range(3, 23):
                self.sh_system.cell(i + 2, k).value = 0
        self.excel_files['system'] = os.path.join(target_dir, FILE_SYSTEM)
        self.save_to_excel('system')

    def _init_user_sheet(self, target_dir):
        for i in range(self.ev[0]):
            self.sh_user.cell(i + 2, 1).value = f'user{i + 1}'
            self.set_reorg_time(i, 0)
            self.set_invalid_refund_available(i, 0)
            self.set_total_payment_specific_user(i, self.ev[0])
            self.set_payable(i, 'yes')
            self.set_defect_count(i, 0)
        self.excel_files['user'] = os.path.join(target_dir, FILE_USER)
        self.save_to_excel('user')

    def _checksum(self, syfunc: int):
        checked_vals = []
        for i in range(self.ev[0]):
            cur_subgroup = self.get_cur_subgroup(i)
            if cur_subgroup == 0 or self.get_cur_status(i) == 'defected':
                continue
            cur_remaining = self.get_remaining_num_cur_subgroup(i)
            if cur_subgroup not in checked_vals:
                remainings = [self.get_remaining_num_cur_subgroup(j) for j in range(self.ev[0])
                              if self.get_cur_subgroup(j) == cur_subgroup and self.get_cur_status(j) != 'defected']
                if len(set(remainings)) > 1:
                    msg = f'______________ Period {self.counter} :: SyFunc {syfunc} _checksum failed(i={i}): ' \
                          f'Remaining numbers in a same subgroup({cur_subgroup}) must be identical, but - {remainings}'
                    logger.error(msg)
                if len(remainings) != cur_remaining:
                    msg = f'______________ Period {self.counter} :: SyFunc {syfunc} _checksum failed(i={i}): ' \
                          f'UsRec4 value({cur_remaining}) doesn\'t match with {len(remainings)}'
                    logger.error(msg)
                checked_vals.append(cur_subgroup)

    def _checksum_sr1(self, _sy_rec1_val: int, syfunc: int):
        counter = len([i for i in range(self.ev[0]) if self.get_cur_subgroup(i) == 0])
        if self.ev[0] - _sy_rec1_val != counter:
            msg = f'______________ Period {self.counter} SyFunc {syfunc} _checksum_sr1 failed: counter = {counter} - ' \
                  f'supposed to be {self.ev[0] - _sy_rec1_val}'
            logger.error(msg)

    def assign_variables(self):
        for i in range(1, 21):
            self.sy_rec_p[i] = self.sh_system.cell(self.counter * 3 - 1, i + 2)
            self.sy_rec_f[i] = self.sh_system.cell(self.counter * 3, i + 2)
            self.sy_rec_r[i] = self.sh_system.cell(self.counter * 3 + 1, i + 2)

    def save_to_excel(self, db_type):
        if not self.matrix:
            self.wb[db_type].save(self.excel_files[db_type])

    def start_simulation(self, target_dir=RESULT_DIR, count=10):
        s_time = time.time()
        target_dir = os.path.join(target_dir, datetime.now().strftime('%m_%d_%Y__%H_%M_%S'))
        os.makedirs(target_dir, exist_ok=True)
        self._init_user_sheet(target_dir)
        self._init_system_sheet(target_dir)

        logger.debug(f'EV1(Total members in the group): {self.ev[0]}')

        # Subgroup  # FUNCTION FOR SUBGROUP EXECUTION
        step1_ev1 = self.ev[0]
        step2 = self.ev[0] / 5
        step3 = round(step2 / 2.3333)
        step4 = step3 * 5
        step5 = step1_ev1 - step4
        step6 = step5 / 6
        step7 = round(step6 / 2)
        step8 = step7 * 6
        step9 = step5 - step8
        step10 = step9 / 7
        step11 = int(step10 / 2)
        step12 = step11 * 7
        step13 = step9 - step12
        step14 = int(step13 / 4)
        step15 = step13 % 4
        if step15 == 1:
            step3 -= 1
            step7 += 1
        elif step15 == 2:
            step3 -= 1
            step11 += 1
        elif step15 == 3:
            step3 -= 1
            step14 += 2

        # Assigning number to the group, condition checking for group == 4
        group_num = 1
        group_mem_count = 0
        temp_val_four = step14 * 4
        offset = 0
        for i in range(temp_val_four):
            self.set_orig_subgroup(i + offset, group_num)
            self.set_remaining_num_orig_subgroup(i + offset, 4)
            self.set_cur_subgroup(i + offset, group_num)
            self.set_remaining_num_cur_subgroup(i + offset, 4)
            group_mem_count += 1
            if group_mem_count == 4:
                group_num += 1
                group_mem_count = 0
        offset += temp_val_four

        # condition checking for group == 5
        temp_val_five = step3 * 5
        for i in range(temp_val_five):
            self.set_orig_subgroup(i + offset, group_num)
            self.set_remaining_num_orig_subgroup(i + offset, 5)
            self.set_cur_subgroup(i + offset, group_num)
            self.set_remaining_num_cur_subgroup(i + offset, 5)
            group_mem_count += 1
            if group_mem_count == 5:
                group_num += 1
                group_mem_count = 0
        offset += temp_val_five

        # condition checking for group == 6
        temp_val_six = step7 * 6
        for i in range(temp_val_six):
            self.set_orig_subgroup(i + offset, group_num)
            self.set_remaining_num_orig_subgroup(i + offset, 6)
            self.set_cur_subgroup(i + offset, group_num)
            self.set_remaining_num_cur_subgroup(i + offset, 6)
            group_mem_count += 1
            if group_mem_count == 6:
                group_num += 1
                group_mem_count = 0
        offset += temp_val_six

        # condition checking for group == 7
        temp_val_seven = step11 * 7
        for i in range(temp_val_seven):
            self.set_orig_subgroup(i + offset, group_num)
            self.set_remaining_num_orig_subgroup(i + offset, 7)
            self.set_cur_subgroup(i + offset, group_num)
            self.set_remaining_num_cur_subgroup(i + offset, 7)
            group_mem_count += 1
            if group_mem_count == 7:
                group_num += 1
                group_mem_count = 0

        checksum = offset + temp_val_seven
        if checksum != self.ev[0]:
            raise ValueError(f"Initial group checksum failed: checksum:{checksum} != EV1:{self.ev[0]}")
        logger.debug({"D": temp_val_four, "A": temp_val_five, "B": temp_val_six, "C": temp_val_seven})

        # setting valid to UsRec5
        for i in range(self.ev[0]):
            self.set_subgroup_status(i, 'valid')
        logger.debug(
            f'Group4 members: {step14}, Group5 members: {step3}, Group6 members: {step7}, '
            f'Group7 members: {step11}, Total group: {step14 * 4 + step3 * 5 + step7 * 6 + step11 * 7}')

        # ROLE 1
        role_ev4 = int(self.ev[0] * self.ev[3])     # EV 4 = Percentage of honest defectors
        defectors = random.sample(range(self.ev[0]), role_ev4)
        non_defectors = [i for i in range(self.ev[0]) if i not in defectors]
        # EV 5 = Percentage of low-morale members
        role_ev5 = int(self.ev[0] * self.ev[4])
        low_morale_list = random.sample(non_defectors, role_ev5)
        for i in range(self.ev[0]):
            self.set_primary_role(
                i, 'defector' if i in defectors else 'low-morale' if i in low_morale_list else 'unity-role')

        # ROLE 2
        # temp_val_four users and pick remaining users randomly to be equal with EV6
        # (percentage of members who are unwilling to act alone)
        remaining_pct = int(self.ev[5] * self.ev[0]) - temp_val_four
        if remaining_pct > 0:
            rand_dep_user = random.sample(range(temp_val_four, self.ev[0]), remaining_pct)
        else:
            rand_dep_user = []
        for i in range(self.ev[0]):
            self.set_secondary_role(
                i, 'dependent' if (i < temp_val_four or i in rand_dep_user) else 'independent')

        for self.counter in range(1, count + 1):
            # logger.info(f'Current period is: {self.counter}')
            self.assign_variables()
            if self.counter == 1:
                self.user_func_1()
            else:
                self.user_func_2()

            self.sys_func_3()

            self.sys_func_4()

            # ___SyFunc5___
            if self.counter == 1:
                self.sy_rec_f[9].value = self.sy_rec_f[3].value * self.sy_rec_f[19].value
                self.sh_system.cell(4, 11).value = self.sy_rec_f[9].value

            self.sys_func_6()

            self.sys_func_7()
            self._checksum(syfunc=7)       # SysFunc7 doesn't perform checksum inside it.

            self.sys_func_8()

            self.sys_func_9()

            # ___SyFunc11___  (Reorg Stage 7)
            total = self.sy_rec_r[3].value + self.sy_rec_r[5].value + self.sy_rec_r[7].value
            if self.counter != count and total > 0:
                # copy values of previous to current
                sy_rec_new_p = [self.sh_system.cell(2, k + 1) for k in range(23)]
                for k in range(1, 21):
                    sy_rec_new_p[k] = self.sh_system.cell(self.counter * 3 + 2, k + 2)
                    sy_rec_new_p[k].value = self.sy_rec_r[k].value

                # Overwriting values in new row
                sy_rec_new_p[18].value = sy_rec_new_p[17].value
                for k in {3, 5, 6, 9, 11, 13, 14, 15, 17}:
                    sy_rec_new_p[k].value = 0

                self._checksum(11)
                self._checksum_sr1(sy_rec_new_p[1].value, 11)
            else:
                self.save_to_excel('user')
                self.save_to_excel('system')
                logger.info(f'Complete at period {self.counter}, elapsed: {time.time() - s_time}')
                percent = round(self.sh_system.cell(3, 5).value / self.ev[0] * 100, 2)
                inc_premium = round((self.sy_rec_f[19].value / self.sh_system.cell(2, 21).value) * 100, 2)
                result_file = os.path.join(target_dir, "result.txt")
                results = [
                    self.ev[0],
                    self.sy_rec_r[1].value,
                    round(((self.ev[0] - self.sy_rec_r[1].value) / self.ev[0]) * 100, 2),
                    round(self.sh_system.cell(2, 21).value),
                    int(self.sy_rec_f[19].value),
                    inc_premium,
                    self.sh_system.cell(3, 5).value,
                    self.ev[3] * 100,
                    percent,
                    self.pv[4] * 100
                ]
                lines = [
                    f'{results[0]} is the number of members at the start of the simulation\n',
                    f'{results[1]} is the number of valid members remaining at the end '
                    f'of the simulation\n',
                    f'{results[2]}% of '
                    f'policyholders left the group by end of simulation\n',
                    f'{results[3]} was the initial premium members were '
                    f'asked to pay.\n',
                    f'{results[4]} is the final premium members were asked to pay.\n',
                    f'Premiums increased by {results[5]}% by end of simulation\n',
                    f'self.SyRec 3 (period 0 finalize) = {results[6]}\n',
                    f'{results[7]}% of policyholders who were assigned to defect\n',
                    f'{results[8]}% of policyholders who actually defected\n',
                    f'{results[9]}% was the initial collapse threshold set for PV 5\n'
                ]
                logger.info('\n' + ''.join(lines))
                if not self.matrix:
                    with open(result_file, 'w') as f:
                        f.writelines(lines)
                else:
                    shutil.rmtree(target_dir, ignore_errors=True)
                return results

    def user_func_1(self):
        """
        User defection function
        """
        # Path 1
        for i in range(self.ev[0]):
            if self.get_primary_role(i) == 'defector' and self.get_secondary_role(i) == 'dependent':
                # Increase the defect counter of all subgroup members where the current user is involved.
                cur_subgroup = self.get_cur_subgroup(i)
                for j in range(self.ev[0]):
                    if self.get_primary_role(j) == 'defector' and self.get_secondary_role(j) == 'dependent' \
                            and self.get_cur_subgroup(j) == cur_subgroup:
                        # Increase the defect count
                        self.set_defect_count(j, self.get_defect_count(j) + 1)

        for i in range(self.ev[0]):
            if self.get_primary_role(i) == 'defector':
                if self.get_defect_count(i) >= self.ev[6] or self.get_secondary_role(i) == 'independent':  # Path 2
                    self.sy_rec_p[1].value -= 1  # Decrease valid members remaining
                    self.sy_rec_p[3].value += 1  # Increase members defected
                    self.sy_rec_p[5].value += 1  # Increase members skipped
                    self._poison_a_user(i)
                if self.get_defect_count(i) < self.ev[6] and self.get_secondary_role(i) == 'dependent':  # Path 3
                    self.set_primary_role(i, 'low-morale')
        self._checksum(syfunc=1)

    def user_func_2(self):
        """"
        Pay Stage 2,  User skip function
        """
        slope = (self.pv[3] - self.pv[1]) / (self.pv[2] - self.pv[0])
        cur_total_payment = float(self.sy_rec_p[19].value)
        prev_total_payment = float(self.sh_system.cell(self.counter * 3 - 1 - 3, 21).value)
        if prev_total_payment > 0:
            inc_premium = max((cur_total_payment / prev_total_payment) - 1, 0)
        else:
            inc_premium = 0

        valid_users = [i for i in range(self.ev[0]) if self.get_subgroup_status(i) == 'valid']
        skip_count = 0
        if inc_premium >= self.pv[0]:  # PATH1
            skip_percent = slope * (inc_premium - self.pv[0]) + self.pv[1]
            skip_count = round(self.sy_rec_p[1].value * skip_percent)
        else:
            if cur_total_payment / (self.ev[9] / self.ev[0]) - 1 >= self.pv[4]:  # PATH2
                skip_count = round(self.sy_rec_p[1].value * self.pv[5])
            else:  # PATH3
                if self.ev[7] != 0:
                    self.ev[7] -= 1
                    skip_count = 1
        if valid_users and skip_count > 0:
            skip_users = random.sample(valid_users, min(skip_count, len(valid_users)))
            for i in skip_users:
                self.set_payable(i, 'no')

    def _poison_a_user(self, index):
        cur_subgroup = self.get_cur_subgroup(index)
        orig_subgroup = self.get_orig_subgroup(index)
        for j in range(self.ev[0]):
            if self.get_cur_subgroup(j) == cur_subgroup:
                # Decrease the number count of the current subgroup
                self.set_remaining_num_cur_subgroup(j, self.get_remaining_num_cur_subgroup(j) - 1)
                if self.get_orig_subgroup(j) == orig_subgroup:
                    # Decrease the number count of original subgroup
                    self.set_remaining_num_orig_subgroup(j, self.get_remaining_num_orig_subgroup(j) - 1)
        self.set_cur_subgroup(index, 0)
        self.set_remaining_num_cur_subgroup(index, 0)
        self.set_subgroup_status(index, 'NR')
        self.set_cur_status(index, 'NR')
        self.set_payable(index, 'NR')

    def sys_func_3(self):
        """"
        Pay Stage 3, Validate premium function
        """
        valid_users = [i for i in range(self.ev[0]) if self.get_subgroup_status(i) == 'valid']

        for i in valid_users:
            if self.get_payable(i) == 'no':
                self.set_cur_status(i, 'skipped')
                self.sy_rec_p[1].value -= 1  # Decrease valid members remaining
                self.sy_rec_p[5].value += 1  # Increase members skipped
                self._poison_a_user(i)
            elif self.get_payable(i) == 'yes':
                self.set_cur_status(i, 'paid')

        self._checksum(3)
        self._checksum_sr1(self.sy_rec_p[1].value, 3)

    def sys_func_4(self):
        """"
        Pay Stage 4, Invalidate subgroup function
        """
        for i in range(self.ev[0]):
            if self.get_remaining_num_cur_subgroup(i) in {1, 2, 3} and self.get_cur_status(i) == 'paid':
                self.set_cur_status(i, 'paid-invalid')
                self.set_subgroup_status(i, 'invalid')
                self.set_invalid_refund_available(i, self.get_total_payment_specific_user(i))  # UsRec 10 = UsRec 11
                self.sy_rec_p[6].value += 1  # Increase invalid members count
        # PATH 1
        for k in range(1, 21):
            self.sy_rec_f[k].value = self.sy_rec_p[k].value
            self.sy_rec_r[k].value = self.sy_rec_p[k].value

    def sys_func_6(self):
        """"
        Reorg Stage 1
        """
        invalid_users = [i for i in range(self.ev[0]) if self.get_cur_status(i) == 'paid-invalid']
        for i in invalid_users:
            if (self.get_primary_role(i) == 'low-morale' and random.uniform(0, 1) < self.ev[8]) or \
                    (self.get_secondary_role(i) == 'dependent' and self.get_remaining_num_orig_subgroup(i) < 2):
                self.set_cur_status(i, 'quit')
                self.sy_rec_r[1].value -= 1
                self.sy_rec_r[7].value += 1
                self._poison_a_user(i)
            else:  # PATH 4
                self.sy_rec_r[8].value += 1

        self._checksum(6)
        self._checksum_sr1(self.sy_rec_r[1].value, 6)

    def sys_func_7(self):
        """"
        Reorg Stage 2
        """
        invalid_users = [i for i in range(self.ev[0]) if self.get_cur_status(i) == 'paid-invalid']
        for path in {1, 2}:
            path_users = [i for i in invalid_users if self.get_remaining_num_cur_subgroup(i) == path]
            # First Attempt
            invalid_list = list(set([self.get_cur_subgroup(i) for i in path_users]))
            valid_list = list(set(
                [self.get_cur_subgroup(i) for i in range(self.ev[0])
                 if self.get_subgroup_status(i) == 'valid' and self.get_remaining_num_cur_subgroup(i) == (7 - path)]))
            while True:
                # Assignment First Attempt
                if invalid_list and valid_list:
                    need_match = invalid_list[0]
                    give_match = random.sample(valid_list, 1)[0]
                    for i in path_users[:]:
                        if self.get_cur_subgroup(i) == need_match:
                            self.set_cur_subgroup(i, give_match)
                            self.set_remaining_num_cur_subgroup(i, 7)
                            self.set_subgroup_status(i, 'valid')
                            self.set_cur_status(i, 'reorg')
                            self.set_reorg_time(i, self.get_reorg_time(i) + 1)
                            path_users.remove(i)
                    invalid_list.remove(need_match)
                    for i in range(self.ev[0]):
                        if self.get_cur_subgroup(i) == give_match:
                            self.set_remaining_num_cur_subgroup(i, 7)
                    valid_list.remove(give_match)
                if not invalid_list:
                    if path_users:
                        logger.warning(f"Period {self.counter}, SysFunc7: Path{path} invalid is empty "
                                       f"but run set is not empty in the 1st attempt!")
                    break
                elif not valid_list:      # Second attempt
                    filtered_list = list(set(
                        [self.get_cur_subgroup(i) for i in range(self.ev[0])
                         if self.get_subgroup_status(i) == 'valid' and
                            self.get_remaining_num_cur_subgroup(i) == (6 - path)]))
                    while filtered_list:
                        need_match = invalid_list[0]
                        give_match = random.sample(filtered_list, 1)[0]
                        for i in path_users[:]:
                            if self.get_cur_subgroup(i) == need_match:
                                self.set_cur_subgroup(i, give_match)
                                self.set_remaining_num_cur_subgroup(i, 6)
                                self.set_subgroup_status(i, 'valid')
                                self.set_cur_status(i, 'reorg')
                                self.set_reorg_time(i, self.get_reorg_time(i) + 1)
                                path_users.remove(i)
                        invalid_list.remove(need_match)
                        for i in range(self.ev[0]):
                            if self.get_cur_subgroup(i) == give_match:
                                self.set_remaining_num_cur_subgroup(i, 6)
                        filtered_list.remove(give_match)
                        if not invalid_list:
                            if path_users:
                                logger.warning(f"Period {self.counter}, SysFunc7: Path{path} invalid is empty "
                                               f"but run set is not empty in the 2nd attempt!")
                            break
                    break

        # Path 3
        path_3_users = [i for i in invalid_users if self.get_remaining_num_cur_subgroup(i) == 3]
        invalid_list = list(set([self.get_cur_subgroup(i) for i in path_3_users]))
        while len(invalid_list) >= 2:      # Path 3 Assignment first attempt
            need_match = invalid_list[0]
            give_match = invalid_list[1]
            for i in path_3_users:
                if self.get_cur_subgroup == need_match:
                    self.set_cur_subgroup(i, give_match)
                    self.set_remaining_num_cur_subgroup(i, 6)
                    self.set_subgroup_status(i, 'valid')
                    self.set_cur_status(i, 'reorg')
                    self.set_reorg_time(i, self.get_reorg_time(i) + 1)
            invalid_list.remove(need_match)
            for i in range(self.ev[0]):
                if self.get_cur_subgroup(i) == give_match and i not in path_3_users:
                    self.set_remaining_num_cur_subgroup(i, 6)
                    self.set_subgroup_status(i, 'valid')
                    self.set_cur_status(i, 'reorg')
                    self.set_reorg_time(i, self.get_reorg_time(i) + 1)
            invalid_list.remove(give_match)
            path_3_users = [i for i in path_3_users if self.get_cur_subgroup(i) not in {need_match, give_match}]
            if not invalid_list:
                if not path_3_users:
                    return
        # Path 3 Second Attempt
        valid_list = list(set(
            [self.get_cur_subgroup(i) for i in range(self.ev[0])
             if self.get_subgroup_status(i) == 'valid' and self.get_remaining_num_cur_subgroup(i) == 4]))
        # TODO: If invalid_list is empty and valid_list is not empty?
        if invalid_list and valid_list:
            need_match = invalid_list[0]
            give_match = random.sample(valid_list, 1)[0]
            for i in path_3_users[:]:
                if self.get_cur_subgroup(i) == need_match:
                    self.set_cur_subgroup(i, give_match)
                    self.set_remaining_num_cur_subgroup(i, 7)
                    self.set_subgroup_status(i, 'valid')
                    self.set_cur_status(i, 'reorg')
                    self.set_reorg_time(i, self.get_reorg_time(i) + 1)
                    path_3_users.remove(i)
            for i in range(self.ev[0]):
                if self.get_cur_subgroup(i) == give_match:
                    self.set_remaining_num_cur_subgroup(i, 7)
            valid_list.remove(give_match)
            if path_3_users:
                logger.warning(f"Period {self.counter}, SysFunc7: Path3 invalid is empty but "
                               f"run set is not empty in the 2nd attempt!")

    def sys_func_8(self):
        """"
        Reorg Stage 4
        """
        if self.ev[2] > random.uniform(0, 1):
            self.sy_rec_r[16].value = 'yes'
        else:
            self.sy_rec_r[16].value = "no"
            self.sy_rec_r[17].value = self.sy_rec_r[2].value

        # ___SyFunc8.5___       Reorg Stage 4.5
        self.sy_rec_r[11].value = self.sy_rec_r[5].value * self.sy_rec_r[19].value
        self.sy_rec_r[13].value = self.sy_rec_r[6].value * self.sy_rec_r[19].value

    def sys_func_9(self):
        """"
        Reorg Stage 5
        """
        if self.sy_rec_r[1].value > 0:
            self.sy_rec_r[2].value = self.ev[9] / self.sy_rec_r[1].value
        self.sy_rec_r[14].value = self.sy_rec_r[9].value + self.sy_rec_r[11].value + self.sy_rec_r[13].value
        if self.sy_rec_r[1].value > 0:
            self.sy_rec_r[15].value = self.sy_rec_r[14].value / self.sy_rec_r[1].value
        for i in range(self.ev[0]):
            invalid_refund = self.get_invalid_refund_available(i)
            if invalid_refund != 0:
                self.set_total_payment_specific_user(
                    i, self.sy_rec_r[2].value + self.sy_rec_r[15].value - invalid_refund)
                self.set_invalid_refund_available(i, 0)
            else:
                sr18 = self.sy_rec_r[18].value
                self.set_total_payment_specific_user(
                    i, self.sy_rec_r[2].value + self.sy_rec_r[15].value - (sr18 if sr18 is not None else 0))
        self.sy_rec_r[19].value = self.sy_rec_r[2].value + self.sy_rec_r[15].value

    # ============   User Rec Functions   ===========================

    def get_orig_subgroup(self, index):
        # UsRec1
        return self.sh_user.cell(index + 2, 2).value

    def set_orig_subgroup(self, index, val):
        # UsRec1
        self.sh_user.cell(index + 2, 2).value = val

    def get_remaining_num_orig_subgroup(self, index):
        # UsRec2
        return self.sh_user.cell(index + 2, 3).value

    def set_remaining_num_orig_subgroup(self, index, num):
        # UsRec2
        self.sh_user.cell(index + 2, 3).value = max(num, 0)

    def get_cur_subgroup(self, index):
        # UsRec3
        return self.sh_user.cell(index + 2, 4).value

    def set_cur_subgroup(self, index, num):
        # UsRec3
        self.sh_user.cell(index + 2, 4).value = num

    def get_remaining_num_cur_subgroup(self, index):
        # UsRec4
        return self.sh_user.cell(index + 2, 5).value

    def set_remaining_num_cur_subgroup(self, index, num):
        # UsRec4
        self.sh_user.cell(index + 2, 5).value = max(num, 0)

    def get_subgroup_status(self, index):
        # UsRec5
        return self.sh_user.cell(index + 2, 6).value

    def set_subgroup_status(self, index, state):
        # UsRec5
        self.sh_user.cell(index + 2, 6).value = state

    def get_primary_role(self, index):
        # UsRec6
        return self.sh_user.cell(index + 2, 7).value

    def set_primary_role(self, index, role):
        # UsRec6
        self.sh_user.cell(index + 2, 7).value = role

    def get_secondary_role(self, index):
        # UsRec7
        return self.sh_user.cell(index + 2, 8).value

    def set_secondary_role(self, index, val):
        # UsRec7
        self.sh_user.cell(index + 2, 8).value = val

    def get_cur_status(self, index):
        # UsRec8
        return self.sh_user.cell(index + 2, 9).value

    def set_cur_status(self, index, state):
        # UsRec8
        self.sh_user.cell(index + 2, 9).value = state

    def get_reorg_time(self, index):
        # UsRec9
        return self.sh_user.cell(index + 2, 10).value

    def set_reorg_time(self, index, val):
        # UsRec9
        self.sh_user.cell(index + 2, 10).value = val

    def get_invalid_refund_available(self, index):
        # UsRec10
        return self.sh_user.cell(index + 2, 11).value

    def set_invalid_refund_available(self, index, val):
        # UsRec10
        self.sh_user.cell(index + 2, 11).value = val

    def get_total_payment_specific_user(self, index):
        # UsRec11
        return self.sh_user.cell(index + 2, 12).value

    def set_total_payment_specific_user(self, index, val):
        # UsRec11
        self.sh_user.cell(index + 2, 12).value = val

    def get_payable(self, index):
        # UsRec12
        return self.sh_user.cell(index + 2, 13).value

    def set_payable(self, index, state):
        # UsRec12
        self.sh_user.cell(index + 2, 13).value = state

    def get_defect_count(self, index):
        # UsRec13
        return self.sh_user.cell(index + 2, 14).value

    def set_defect_count(self, index, num):
        # UsRec13
        self.sh_user.cell(index + 2, 14).value = num


if __name__ == '__main__':

    _ev = [0, 0, 0, 0, 0, 0, 0, 0, 0., 0.]
    _pv = []
    tps = TandaPaySimulatorV2(ev=_ev, pv=_pv)
    tps.start_simulation()
